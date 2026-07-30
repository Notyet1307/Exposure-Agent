#!/usr/bin/env python3
"""Reproducible XLSX parser and ZIP-resource boundary probe for issue #33."""

from __future__ import annotations

import argparse
import json
import re
import struct
import sys
import tempfile
import time
import tracemalloc
import zlib
from collections.abc import Iterable, Sequence
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from pathlib import Path, PurePosixPath
from typing import Any, Final
from zipfile import ZIP_DEFLATED, ZIP_STORED, ZipFile, ZipInfo

import defusedxml  # type: ignore[import-untyped]
import openpyxl  # type: ignore[import-untyped]
from defusedxml import ElementTree
from defusedxml.common import DefusedXmlException  # type: ignore[import-untyped]
from openpyxl import Workbook, load_workbook
from openpyxl.xml.functions import (  # type: ignore[import-untyped]
    DEFUSEDXML as OPENPYXL_DEFUSEDXML,
)

REQUEST_LIMIT_BYTES: Final = 20 * 1024 * 1024
MAX_ZIP_ENTRIES: Final = 2_048
MAX_ENTRY_UNCOMPRESSED_BYTES: Final = 64 * 1024 * 1024
MAX_TOTAL_UNCOMPRESSED_BYTES: Final = 256 * 1024 * 1024
MAX_CENTRAL_DIRECTORY_BYTES: Final = 4 * 1024 * 1024
MAX_EOCD_TAIL_BYTES: Final = 65_557
ZIP64_SENTINEL_16: Final = 0xFFFF
ZIP64_SENTINEL_32: Final = 0xFFFFFFFF
FIXED_TIME: Final = datetime(2026, 1, 1, tzinfo=UTC)

EOCD = struct.Struct("<4s4H2LH")
CENTRAL_HEADER = struct.Struct("<4s6H3L5H2L")
EOCD_SIGNATURE: Final = b"PK\x05\x06"
CENTRAL_SIGNATURE: Final = b"PK\x01\x02"
REL_NS: Final = "http://schemas.openxmlformats.org/package/2006/relationships"
MAIN_NS: Final = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS: Final = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)

NORMAL_FIXTURES: Final = (
    "default_v1",
    "near_request_limit",
    "row_shared_style_boundary",
)
RESOURCE_FIXTURES: Final = (
    "entry_count_bomb",
    "compression_bomb",
    "total_size_bomb",
)
FORBIDDEN_FIXTURES: Final = (
    "formula",
    "external_link",
    "data_connection",
    "hidden_sheet",
    "embedded_object",
)
ALL_FIXTURES: Final = NORMAL_FIXTURES + RESOURCE_FIXTURES + FORBIDDEN_FIXTURES


@dataclass(frozen=True)
class ZipStats:
    entries: int
    central_directory_bytes: int
    compressed_bytes: int
    total_uncompressed_bytes: int
    largest_entry_uncompressed_bytes: int
    largest_entry_name: str


@dataclass(frozen=True)
class InspectionResult:
    status: str
    rows: int
    elapsed_seconds: float
    peak_memory_bytes: int
    zip_stats: ZipStats


class WorkbookRejected(Exception):
    """A stable, redacted workbook rejection from the investigation seam."""

    def __init__(
        self,
        category: str,
        reason: str,
        phase: str,
        *,
        zip_stats: ZipStats | None = None,
    ) -> None:
        super().__init__(category)
        self.category = category
        self.reason = reason
        self.phase = phase
        self.zip_stats = zip_stats


def _resource_rejection(reason: str, stats: ZipStats | None = None) -> WorkbookRejected:
    return WorkbookRejected(
        "workbook_resource_limit", reason, "central_directory", zip_stats=stats
    )


def _safe_member_name(raw_name: bytes, flags: int) -> str:
    encoding = "utf-8" if flags & 0x800 else "cp437"
    try:
        name = raw_name.decode(encoding)
    except UnicodeDecodeError as exc:
        raise WorkbookRejected(
            "malformed_workbook", "invalid_zip_member_name", "central_directory"
        ) from exc
    pure_name = PurePosixPath(name)
    if not name or "\\" in name or pure_name.is_absolute() or ".." in pure_name.parts:
        raise WorkbookRejected(
            "malformed_workbook", "unsafe_zip_member_name", "central_directory"
        )
    return name


def read_bounded_central_directory(path: Path) -> ZipStats:
    """Read the EOCD tail and bounded central directory without opening ZIP entries."""
    file_size = path.stat().st_size
    if file_size < EOCD.size:
        raise WorkbookRejected("malformed_workbook", "missing_eocd", "eocd")

    tail_size = min(file_size, MAX_EOCD_TAIL_BYTES)
    with path.open("rb") as workbook_file:
        workbook_file.seek(file_size - tail_size)
        tail = workbook_file.read(tail_size)
        signature_offset = tail.rfind(EOCD_SIGNATURE)
        if signature_offset < 0 or signature_offset + EOCD.size > len(tail):
            raise WorkbookRejected("malformed_workbook", "missing_eocd", "eocd")
        (
            signature,
            disk_number,
            central_disk,
            disk_entries,
            total_entries,
            central_size,
            central_offset,
            comment_length,
        ) = EOCD.unpack_from(tail, signature_offset)
        eocd_offset = file_size - tail_size + signature_offset
        if (
            signature != EOCD_SIGNATURE
            or eocd_offset + EOCD.size + comment_length != file_size
        ):
            raise WorkbookRejected("malformed_workbook", "invalid_eocd", "eocd")
        if disk_number != 0 or central_disk != 0 or disk_entries != total_entries:
            raise WorkbookRejected("malformed_workbook", "multi_disk_zip", "eocd")
        if total_entries == ZIP64_SENTINEL_16 or central_size == ZIP64_SENTINEL_32:
            raise WorkbookRejected("malformed_workbook", "zip64_not_supported", "eocd")
        if total_entries > MAX_ZIP_ENTRIES:
            raise _resource_rejection("zip_entry_count")
        if central_size > MAX_CENTRAL_DIRECTORY_BYTES:
            raise _resource_rejection("central_directory_size")
        if central_offset + central_size > eocd_offset:
            raise WorkbookRejected(
                "malformed_workbook", "invalid_central_directory_bounds", "eocd"
            )

        workbook_file.seek(central_offset)
        central = workbook_file.read(central_size)

    cursor = 0
    compressed_total = 0
    uncompressed_total = 0
    largest_size = 0
    largest_name = ""
    names: set[str] = set()
    parsed_entries = 0
    while cursor < len(central):
        if cursor + CENTRAL_HEADER.size > len(central):
            raise WorkbookRejected(
                "malformed_workbook", "truncated_central_directory", "central_directory"
            )
        values = CENTRAL_HEADER.unpack_from(central, cursor)
        (
            signature,
            _made_by,
            _needed,
            flags,
            compression,
            _modified_time,
            _modified_date,
            _crc32,
            compressed_size,
            uncompressed_size,
            name_length,
            extra_length,
            member_comment_length,
            disk_start,
            _internal_attributes,
            _external_attributes,
            local_offset,
        ) = values
        if signature != CENTRAL_SIGNATURE:
            raise WorkbookRejected(
                "malformed_workbook", "invalid_central_directory", "central_directory"
            )
        variable_size = name_length + extra_length + member_comment_length
        next_cursor = cursor + CENTRAL_HEADER.size + variable_size
        if next_cursor > len(central):
            raise WorkbookRejected(
                "malformed_workbook", "truncated_central_directory", "central_directory"
            )
        if (
            compressed_size == ZIP64_SENTINEL_32
            or uncompressed_size == ZIP64_SENTINEL_32
            or local_offset == ZIP64_SENTINEL_32
            or disk_start == ZIP64_SENTINEL_16
        ):
            raise WorkbookRejected(
                "malformed_workbook", "zip64_not_supported", "central_directory"
            )
        if flags & 0x1:
            raise WorkbookRejected(
                "malformed_workbook", "encrypted_zip_member", "central_directory"
            )
        if compression not in (ZIP_STORED, ZIP_DEFLATED):
            raise WorkbookRejected(
                "malformed_workbook", "unsupported_zip_compression", "central_directory"
            )
        name_start = cursor + CENTRAL_HEADER.size
        name = _safe_member_name(central[name_start : name_start + name_length], flags)
        if name in names:
            raise WorkbookRejected(
                "malformed_workbook", "duplicate_zip_member", "central_directory"
            )
        names.add(name)
        compressed_total += compressed_size
        uncompressed_total += uncompressed_size
        if uncompressed_size > largest_size:
            largest_size = uncompressed_size
            largest_name = name
        cursor = next_cursor
        parsed_entries += 1

    if parsed_entries != total_entries or cursor != central_size:
        raise WorkbookRejected(
            "malformed_workbook",
            "central_directory_count_mismatch",
            "central_directory",
        )

    stats = ZipStats(
        entries=parsed_entries,
        central_directory_bytes=central_size,
        compressed_bytes=compressed_total,
        total_uncompressed_bytes=uncompressed_total,
        largest_entry_uncompressed_bytes=largest_size,
        largest_entry_name=largest_name,
    )
    if largest_size > MAX_ENTRY_UNCOMPRESSED_BYTES:
        raise _resource_rejection("single_entry_uncompressed_size", stats)
    if uncompressed_total > MAX_TOTAL_UNCOMPRESSED_BYTES:
        raise _resource_rejection("total_uncompressed_size", stats)
    return stats


def _relationship_rejection(relationship_type: str) -> str | None:
    suffix = relationship_type.rsplit("/", maxsplit=1)[-1]
    if suffix in {"externalLink", "externalLinkPath"}:
        return "external_link"
    if suffix == "connections":
        return "data_connection"
    if suffix in {
        "activeXControl",
        "activeXControlBinary",
        "attachedToolbars",
        "control",
        "ctrlProp",
        "oleObject",
        "package",
        "vbaProject",
    }:
        return "embedded_active_object"
    return None


def _content_type_rejection(content_type: str) -> str | None:
    lowered = content_type.lower()
    if "externallink" in lowered:
        return "external_link"
    if "connections+xml" in lowered:
        return "data_connection"
    if any(
        marker in lowered
        for marker in (
            "activex",
            "attachedtoolbars",
            "ctrlprop",
            "oleobject",
            "vbaproject",
        )
    ):
        return "embedded_active_object"
    return None


def _inspect_forbidden_ooxml(archive: ZipFile) -> None:
    names = set(archive.namelist())
    if "[Content_Types].xml" not in names:
        raise WorkbookRejected(
            "malformed_workbook", "missing_content_types", "ooxml_structure"
        )
    try:
        with archive.open("[Content_Types].xml") as content_types_file:
            for _event, element in ElementTree.iterparse(
                content_types_file, events=("end",)
            ):
                reason = _content_type_rejection(element.attrib.get("ContentType", ""))
                if reason is not None:
                    raise WorkbookRejected(
                        "unsupported_workbook_feature",
                        reason,
                        "ooxml_content_types",
                    )
                element.clear()
    except (ElementTree.ParseError, DefusedXmlException) as exc:
        raise WorkbookRejected(
            "malformed_workbook", "invalid_ooxml_xml", "ooxml_structure"
        ) from exc

    for name in names:
        if name.startswith("xl/externalLinks/"):
            raise WorkbookRejected(
                "unsupported_workbook_feature", "external_link", "ooxml_structure"
            )
        if name == "xl/connections.xml":
            raise WorkbookRejected(
                "unsupported_workbook_feature", "data_connection", "ooxml_structure"
            )
        if name.startswith(
            ("xl/activeX/", "xl/ctrlProps/", "xl/embeddings/")
        ) or name in {
            "xl/vbaProject.bin",
            "xl/attachedToolbars.bin",
        }:
            raise WorkbookRejected(
                "unsupported_workbook_feature",
                "embedded_active_object",
                "ooxml_structure",
            )

    try:
        relationship_names = sorted(name for name in names if name.endswith(".rels"))
        for relationship_name in relationship_names:
            with archive.open(relationship_name) as relationship_file:
                for _event, element in ElementTree.iterparse(
                    relationship_file, events=("end",)
                ):
                    if element.tag == f"{{{REL_NS}}}Relationship":
                        reason = _relationship_rejection(element.attrib.get("Type", ""))
                        if reason is not None:
                            raise WorkbookRejected(
                                "unsupported_workbook_feature",
                                reason,
                                "ooxml_relationships",
                            )
                    element.clear()

        xml_names = sorted(name for name in names if name.endswith(".xml"))
        for xml_name in xml_names:
            with archive.open(xml_name) as xml_file:
                for _event, element in ElementTree.iterparse(xml_file, events=("end",)):
                    if element.tag.rsplit("}", maxsplit=1)[-1] == "f":
                        raise WorkbookRejected(
                            "unsupported_workbook_feature", "formula", "ooxml_xml"
                        )
                    element.clear()
    except (ElementTree.ParseError, DefusedXmlException) as exc:
        raise WorkbookRejected(
            "malformed_workbook", "invalid_ooxml_xml", "ooxml_structure"
        ) from exc


def inspect_workbook(path: Path) -> InspectionResult:
    """Apply bounded preflight, forbidden-feature checks, then the final parser."""
    tracemalloc.start()
    started = time.perf_counter()
    workbook: Any | None = None
    try:
        stats = read_bounded_central_directory(path)
        try:
            with ZipFile(path) as archive:
                _inspect_forbidden_ooxml(archive)
        except WorkbookRejected:
            raise
        except Exception as exc:
            raise WorkbookRejected(
                "malformed_workbook", "invalid_zip_container", "ooxml_structure"
            ) from exc

        try:
            workbook = load_workbook(
                path, read_only=True, data_only=False, keep_links=True
            )
            if len(workbook.worksheets) != 1:
                hidden = any(
                    worksheet.sheet_state != "visible"
                    for worksheet in workbook.worksheets
                )
                reason = "hidden_sheet" if hidden else "multiple_worksheets"
                raise WorkbookRejected(
                    "unsupported_workbook_feature", reason, "openpyxl"
                )
            worksheet = workbook.worksheets[0]
            if worksheet.sheet_state != "visible":
                raise WorkbookRejected(
                    "unsupported_workbook_feature", "hidden_sheet", "openpyxl"
                )
            non_empty_rows = 0
            for row in worksheet.iter_rows(values_only=False):
                if any(cell.value is not None for cell in row):
                    non_empty_rows += 1
            rows = max(0, non_empty_rows - 1)
        except WorkbookRejected:
            raise
        except Exception as exc:
            raise WorkbookRejected(
                "malformed_workbook", "parser_rejected_workbook", "openpyxl"
            ) from exc
        elapsed = time.perf_counter() - started
        _, peak_memory = tracemalloc.get_traced_memory()
        return InspectionResult(
            status="accepted",
            rows=rows,
            elapsed_seconds=elapsed,
            peak_memory_bytes=peak_memory,
            zip_stats=stats,
        )
    finally:
        if workbook is not None:
            workbook.close()
        tracemalloc.stop()


def _set_reproducible_properties(workbook: Workbook) -> None:
    workbook.properties.creator = "Exposure-Agent fixture"
    workbook.properties.lastModifiedBy = "Exposure-Agent fixture"
    workbook.properties.created = FIXED_TIME.replace(tzinfo=None)
    workbook.properties.modified = FIXED_TIME.replace(tzinfo=None)


def _save_default_workbook(path: Path, *, formula: bool = False) -> None:
    workbook = Workbook()
    _set_reproducible_properties(workbook)
    worksheet = workbook.active
    worksheet.title = "资产清单"
    worksheet.append(
        [
            "资产IP",
            "起始端口",
            "结束端口",
            "是否web界面",
            "web界面url",
            "服务类型",
            "资产负责人",
            "资产所属部门",
            "端口负责人",
            "部门",
        ]
    )
    worksheet.append(
        [
            "192.0.2.10",
            443,
            443,
            "是",
            "https://fixture.example.invalid",
            "Fixture Service",
            "Example Owner",
            "Example Department",
            "Example Port Owner",
            "Example Operations",
        ]
    )
    worksheet.append(
        [
            "198.51.100.20",
            22,
            22,
            "否",
            None,
            "Fixture SSH",
            "Example Owner",
            "Example Department",
            "Example Port Owner",
            "Example Operations",
        ]
    )
    worksheet.append(
        [
            "203.0.113.30",
            8080,
            8080,
            "否",
            None,
            "Fixture API",
            "Example Owner",
            "Example Department",
            "Example Port Owner",
            "Example Operations",
        ]
    )
    if formula:
        worksheet["K2"] = "=1+1"
    workbook.save(path)
    workbook.close()
    _normalize_zip(path)


def _fixed_zip_info(name: str) -> ZipInfo:
    info = ZipInfo(name, date_time=(2026, 1, 1, 0, 0, 0))
    info.compress_type = ZIP_DEFLATED
    info.external_attr = 0o600 << 16
    return info


def _normalize_zip(path: Path) -> None:
    replacement = path.with_suffix(".normalized.xlsx")
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for source_info in source.infolist():
            source_data = source.read(source_info)
            if source_info.filename == "docProps/core.xml":
                source_data = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2026-01-01T00:00:00Z\g<2>",
                    source_data,
                )
            target.writestr(_fixed_zip_info(source_info.filename), source_data)
    replacement.replace(path)


def _append_declared_part(
    path: Path,
    *,
    name: str,
    data: bytes,
    content_type: str,
    relationships_name: str,
    relationship_type: str,
    relationship_target: str,
) -> None:
    replacement = path.with_suffix(".rewritten.xlsx")
    relationship_written = False
    declaration = (
        f'<Override PartName="/{name}" ContentType="{content_type}"/>'
    ).encode()
    relationship = (
        f'<Relationship Id="rIdFixture" Type="{relationship_type}" '
        f'Target="{relationship_target}"/>'
    ).encode()
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for source_info in source.infolist():
            source_data = source.read(source_info)
            if source_info.filename == "[Content_Types].xml":
                source_data = source_data.replace(
                    b"</Types>", declaration + b"</Types>"
                )
            elif source_info.filename == relationships_name:
                source_data = source_data.replace(
                    b"</Relationships>", relationship + b"</Relationships>"
                )
                relationship_written = True
            target.writestr(_fixed_zip_info(source_info.filename), source_data)
        if not relationship_written:
            target.writestr(
                _fixed_zip_info(relationships_name),
                (
                    f'<Relationships xmlns="{REL_NS}">'.encode()
                    + relationship
                    + b"</Relationships>"
                ),
            )
        target.writestr(_fixed_zip_info(name), data)
    replacement.replace(path)


def _relocate_first_worksheet(path: Path) -> None:
    original_name = "xl/worksheets/sheet1.xml"
    relocated_name = "xl/fixture/worksheet.xml"
    replacement = path.with_suffix(".relocated.xlsx")
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        worksheet_data = source.read(original_name)
        for source_info in source.infolist():
            if source_info.filename == original_name:
                continue
            source_data = source.read(source_info)
            if source_info.filename == "[Content_Types].xml":
                source_data = source_data.replace(
                    b"/xl/worksheets/sheet1.xml", b"/xl/fixture/worksheet.xml"
                )
            elif source_info.filename == "xl/_rels/workbook.xml.rels":
                source_data = source_data.replace(
                    b'Target="/xl/worksheets/sheet1.xml"',
                    b'Target="/xl/fixture/worksheet.xml"',
                ).replace(
                    b'Target="worksheets/sheet1.xml"',
                    b'Target="fixture/worksheet.xml"',
                )
            target.writestr(_fixed_zip_info(source_info.filename), source_data)
        target.writestr(_fixed_zip_info(relocated_name), worksheet_data)
    replacement.replace(path)


def _create_png(width: int = 2_048, height: int = 3_300) -> bytes:
    signature = b"\x89PNG\r\n\x1a\n"

    def chunk(kind: bytes, data: bytes) -> bytes:
        return (
            struct.pack(">L", len(data))
            + kind
            + data
            + struct.pack(">L", zlib.crc32(kind + data) & 0xFFFFFFFF)
        )

    header = struct.pack(">2L5B", width, height, 8, 2, 0, 0, 0)
    pixel = bytes(range(256)) * ((width * 3 + 255) // 256)
    scanline = b"\x00" + pixel[: width * 3]
    image_data = zlib.compress(scanline * height, level=0)
    return (
        signature
        + chunk(b"IHDR", header)
        + chunk(b"IDAT", image_data)
        + chunk(b"IEND", b"")
    )


def _add_near_limit_static_image(path: Path) -> None:
    image = _create_png()
    replacement = path.with_suffix(".image.xlsx")
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for source_info in source.infolist():
            data = source.read(source_info)
            if source_info.filename == "[Content_Types].xml":
                data = data.replace(
                    b"</Types>",
                    b'<Default Extension="png" ContentType="image/png"/></Types>',
                )
            elif source_info.filename == "xl/worksheets/sheet1.xml":
                data = data.replace(
                    b"</worksheet>",
                    (
                        b'<drawing xmlns:r="http://schemas.openxmlformats.org/'
                        b'officeDocument/2006/relationships" r:id="rId1"/>'
                        b"</worksheet>"
                    ),
                )
            target.writestr(_fixed_zip_info(source_info.filename), data)
        target.writestr(
            _fixed_zip_info("xl/worksheets/_rels/sheet1.xml.rels"),
            (
                f'<Relationships xmlns="{REL_NS}"><Relationship Id="rId1" '
                f'Type="{OFFICE_REL_NS}/drawing" Target="../drawings/drawing1.xml"/>'
                "</Relationships>"
            ).encode(),
        )
        target.writestr(
            _fixed_zip_info("xl/drawings/drawing1.xml"),
            (
                '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
                'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><xdr:oneCellAnchor>'
                "<xdr:from><xdr:col>11</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>1</xdr:row>"
                '<xdr:rowOff>0</xdr:rowOff></xdr:from><xdr:ext cx="9525" cy="9525"/>'
                '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="1" name="Boundary fixture"/>'
                "<xdr:cNvPicPr/></xdr:nvPicPr><xdr:blipFill><a:blip "
                f'xmlns:r="{OFFICE_REL_NS}" r:embed="rId1"/><a:stretch><a:fillRect/>'
                '</a:stretch></xdr:blipFill><xdr:spPr><a:prstGeom prst="rect"><a:avLst/>'
                "</a:prstGeom></xdr:spPr></xdr:pic><xdr:clientData/></xdr:oneCellAnchor></xdr:wsDr>"
            ).encode(),
        )
        target.writestr(
            _fixed_zip_info("xl/drawings/_rels/drawing1.xml.rels"),
            (
                f'<Relationships xmlns="{REL_NS}"><Relationship Id="rId1" '
                f'Type="{OFFICE_REL_NS}/image" Target="../media/boundary.png"/>'
                "</Relationships>"
            ).encode(),
        )
        image_info = _fixed_zip_info("xl/media/boundary.png")
        image_info.compress_type = ZIP_STORED
        target.writestr(image_info, image)
    replacement.replace(path)


def _stress_workbook_parts(
    row_count: int = 50_000, style_count: int = 64
) -> dict[str, bytes]:
    headers = [
        "资产IP",
        "起始端口",
        "结束端口",
        "是否web界面",
        "web界面url",
        "服务类型",
        "资产负责人",
        "资产所属部门",
        "端口负责人",
        "部门",
    ]
    addresses = (
        [f"192.0.2.{number}" for number in range(1, 255)]
        + [f"198.51.100.{number}" for number in range(1, 255)]
        + [f"203.0.113.{number}" for number in range(1, 255)]
    )
    ownership = [f"Example Responsibility {number:04d}" for number in range(1_024)]
    shared_strings = headers + addresses + ["否", "Fixture Service"] + ownership
    string_index = {value: index for index, value in enumerate(shared_strings)}
    shared_xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><sst xmlns="{MAIN_NS}" '
        f'count="{row_count * 8 + len(headers)}" uniqueCount="{len(shared_strings)}">'
        + "".join(f"<si><t>{value}</t></si>" for value in shared_strings)
        + "</sst>"
    )
    rows = [
        '<row r="1">'
        + "".join(
            f'<c r="{chr(65 + column)}1" t="s"><v>{string_index[header]}</v></c>'
            for column, header in enumerate(headers)
        )
        + "</row>"
    ]
    no_index = string_index["否"]
    service_index = string_index["Fixture Service"]
    for offset in range(row_count):
        row_number = offset + 2
        style = offset % style_count
        address_index = string_index[addresses[offset % len(addresses)]]
        owner_index = string_index[ownership[offset % len(ownership)]]
        rows.append(
            f'<row r="{row_number}">'
            f'<c r="A{row_number}" s="{style}" t="s"><v>{address_index}</v></c>'
            f'<c r="B{row_number}" s="{style}" t="n"><v>443</v></c>'
            f'<c r="C{row_number}" s="{style}" t="n"><v>443</v></c>'
            f'<c r="D{row_number}" s="{style}" t="s"><v>{no_index}</v></c>'
            f'<c r="F{row_number}" s="{style}" t="s"><v>{service_index}</v></c>'
            f'<c r="G{row_number}" s="{style}" t="s"><v>{owner_index}</v></c>'
            f'<c r="H{row_number}" s="{style}" t="s"><v>{owner_index}</v></c>'
            f'<c r="I{row_number}" s="{style}" t="s"><v>{owner_index}</v></c>'
            f'<c r="J{row_number}" s="{style}" t="s"><v>{owner_index}</v></c>'
            "</row>"
        )
    sheet_xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><worksheet xmlns="{MAIN_NS}">'
        f'<dimension ref="A1:J{row_count + 1}"/><sheetData>{"".join(rows)}</sheetData></worksheet>'
    )
    xfs = "".join(
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
        for _ in range(style_count)
    )
    styles_xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><styleSheet xmlns="{MAIN_NS}">'
        '<fonts count="1"><font><name val="Arial"/><sz val="10"/></font></fonts>'
        '<fills count="2"><fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="gray125"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        f'<cellXfs count="{style_count}">{xfs}</cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        "</styleSheet>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        "</Types>"
    )
    root_rels = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="{REL_NS}">'
        f'<Relationship Id="rId1" Type="{OFFICE_REL_NS}/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    workbook_xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="{MAIN_NS}" '
        f'xmlns:r="{OFFICE_REL_NS}"><sheets><sheet name="资产清单" sheetId="1" r:id="rId1"/>'
        "</sheets></workbook>"
    )
    workbook_rels = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="{REL_NS}">'
        f'<Relationship Id="rId1" Type="{OFFICE_REL_NS}/worksheet" Target="worksheets/sheet1.xml"/>'
        f'<Relationship Id="rId2" Type="{OFFICE_REL_NS}/styles" Target="styles.xml"/>'
        f'<Relationship Id="rId3" Type="{OFFICE_REL_NS}/sharedStrings" Target="sharedStrings.xml"/>'
        "</Relationships>"
    )
    return {
        "[Content_Types].xml": content_types.encode(),
        "_rels/.rels": root_rels.encode(),
        "xl/workbook.xml": workbook_xml.encode(),
        "xl/_rels/workbook.xml.rels": workbook_rels.encode(),
        "xl/styles.xml": styles_xml.encode(),
        "xl/sharedStrings.xml": shared_xml.encode(),
        "xl/worksheets/sheet1.xml": sheet_xml.encode(),
    }


def _write_parts(path: Path, parts: dict[str, bytes]) -> None:
    with ZipFile(path, "w", compression=ZIP_DEFLATED, compresslevel=9) as archive:
        for name, data in parts.items():
            archive.writestr(_fixed_zip_info(name), data)


def _write_repeated_member(member: Any, size: int, byte: bytes) -> None:
    block = byte * (1024 * 1024)
    remaining = size
    while remaining:
        chunk_size = min(remaining, len(block))
        member.write(block[:chunk_size])
        remaining -= chunk_size


def build_fixture(name: str, path: Path) -> Path:
    """Build one deterministic, sanitized fixture at a caller-owned path."""
    path.parent.mkdir(parents=True, exist_ok=True)
    if name == "row_shared_style_boundary":
        _write_parts(path, _stress_workbook_parts())
        return path
    if name == "entry_count_bomb":
        with ZipFile(path, "w", compression=ZIP_DEFLATED, compresslevel=9) as archive:
            for index in range(MAX_ZIP_ENTRIES + 1):
                archive.writestr(_fixed_zip_info(f"fixture/{index:04d}.bin"), b"")
        return path

    _save_default_workbook(path, formula=name in {"formula", "relocated_formula"})
    if name in {"default_v1", "formula"}:
        return path
    if name == "relocated_formula":
        _relocate_first_worksheet(path)
        return path
    if name == "near_request_limit":
        _add_near_limit_static_image(path)
    elif name == "hidden_sheet":
        workbook = load_workbook(path)
        hidden = workbook.create_sheet("Hidden fixture")
        hidden.sheet_state = "hidden"
        _set_reproducible_properties(workbook)
        workbook.save(path)
        workbook.close()
        _normalize_zip(path)
    elif name == "external_link":
        _append_declared_part(
            path,
            name="xl/fixture/external-link.xml",
            data=f'<externalLink xmlns="{MAIN_NS}"/>'.encode(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml",
            relationships_name="xl/_rels/workbook.xml.rels",
            relationship_type=f"{OFFICE_REL_NS}/externalLink",
            relationship_target="fixture/external-link.xml",
        )
    elif name == "data_connection":
        _append_declared_part(
            path,
            name="xl/fixture/connections.xml",
            data=f'<connections xmlns="{MAIN_NS}" count="0"/>'.encode(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml",
            relationships_name="xl/_rels/workbook.xml.rels",
            relationship_type=f"{OFFICE_REL_NS}/connections",
            relationship_target="fixture/connections.xml",
        )
    elif name == "embedded_object":
        _append_declared_part(
            path,
            name="xl/fixture/object.bin",
            data=b"fixture-active-object",
            content_type="application/vnd.openxmlformats-officedocument.oleObject",
            relationships_name="xl/worksheets/_rels/sheet1.xml.rels",
            relationship_type=f"{OFFICE_REL_NS}/oleObject",
            relationship_target="../fixture/object.bin",
        )
    elif name == "compression_bomb":
        with ZipFile(path, "a", compression=ZIP_DEFLATED, compresslevel=9) as archive:
            info = _fixed_zip_info("xl/sharedStrings.xml")
            with archive.open(info, "w", force_zip64=False) as member:
                _write_repeated_member(member, MAX_ENTRY_UNCOMPRESSED_BYTES + 1, b"A")
    elif name == "total_size_bomb":
        member_size = MAX_TOTAL_UNCOMPRESSED_BYTES // 5 + 1
        with ZipFile(path, "a", compression=ZIP_DEFLATED, compresslevel=9) as archive:
            for index in range(5):
                info = _fixed_zip_info(f"fixture/total-{index}.bin")
                with archive.open(info, "w", force_zip64=False) as member:
                    _write_repeated_member(member, member_size, b"B")
    else:
        raise ValueError(f"unknown fixture kind: {name}")
    return path


def _report_rejection(
    name: str, path: Path, rejection: WorkbookRejected
) -> dict[str, Any]:
    result: dict[str, Any] = {
        "fixture": name,
        "request_bytes": path.stat().st_size,
        "status": "rejected",
        "category": rejection.category,
        "reason": rejection.reason,
        "phase": rejection.phase,
    }
    if rejection.zip_stats is not None:
        result["zip"] = asdict(rejection.zip_stats)
    return result


def run_probe(
    *,
    temp_parent: Path | None = None,
    fixture_names: Sequence[str] = ALL_FIXTURES,
) -> dict[str, Any]:
    """Generate fixtures, inspect them, return a path-free JSON-compatible report."""
    fixture_reports: list[dict[str, Any]] = []
    with tempfile.TemporaryDirectory(
        prefix="xlsx-boundary-", dir=temp_parent
    ) as temporary:
        fixture_directory = Path(temporary)
        for fixture_name in fixture_names:
            fixture_path = build_fixture(
                fixture_name, fixture_directory / f"{fixture_name}.xlsx"
            )
            try:
                inspection = inspect_workbook(fixture_path)
                fixture_reports.append(
                    {
                        "fixture": fixture_name,
                        "request_bytes": fixture_path.stat().st_size,
                        "status": inspection.status,
                        "rows": inspection.rows,
                        "elapsed_seconds": round(inspection.elapsed_seconds, 6),
                        "peak_memory_bytes": inspection.peak_memory_bytes,
                        "zip": asdict(inspection.zip_stats),
                    }
                )
            except WorkbookRejected as rejection:
                fixture_reports.append(
                    _report_rejection(fixture_name, fixture_path, rejection)
                )
    return {
        "python": f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
        "parser": {
            "name": "openpyxl",
            "version": openpyxl.__version__,
            "xml_guard": {
                "name": "defusedxml",
                "version": defusedxml.__version__,
                "openpyxl_enabled": OPENPYXL_DEFUSEDXML,
            },
        },
        "limits": {
            "request_bytes": REQUEST_LIMIT_BYTES,
            "zip_entries": MAX_ZIP_ENTRIES,
            "single_entry_uncompressed_bytes": MAX_ENTRY_UNCOMPRESSED_BYTES,
            "total_uncompressed_bytes": MAX_TOTAL_UNCOMPRESSED_BYTES,
            "central_directory_bytes": MAX_CENTRAL_DIRECTORY_BYTES,
        },
        "fixtures": fixture_reports,
    }


def _parse_fixture_names(values: Iterable[str] | None) -> tuple[str, ...]:
    names = tuple(values or ALL_FIXTURES)
    unknown = sorted(set(names) - set(ALL_FIXTURES))
    if unknown:
        raise ValueError(f"unknown fixture kind: {', '.join(unknown)}")
    return names


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--fixture",
        action="append",
        choices=ALL_FIXTURES,
        help="run only this fixture; repeat for more than one",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="write the redacted JSON report to this path as well as stdout",
    )
    arguments = parser.parse_args(argv)
    report = run_probe(fixture_names=_parse_fixture_names(arguments.fixture))
    serialized = json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    if arguments.output is not None:
        arguments.output.write_text(serialized, encoding="utf-8")
    sys.stdout.write(serialized)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
