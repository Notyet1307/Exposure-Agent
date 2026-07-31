from __future__ import annotations

import ipaddress
import re
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Final

import defusedxml  # type: ignore[import-untyped]
import openpyxl  # type: ignore[import-untyped]
from openpyxl import load_workbook
from openpyxl.xml.functions import (  # type: ignore[import-untyped]
    DEFUSEDXML as OPENPYXL_DEFUSEDXML,
)

from app.domain._xlsx_preflight import (
    PreflightError,
    PreflightResult,
    preflight_xlsx,
)
from app.domain.customer_upload_profiles import (
    OPTIONAL_HEADERS,
    REQUIRED_HEADERS,
    WARNING_HEADERS,
)

MAX_WORKBOOK_BYTES: Final = 20 * 1024 * 1024

_REQUIRED_FIELDS: Final = dict(
    zip(
        ("asset_ip", "start_port", "end_port", "is_web", "web_url"),
        REQUIRED_HEADERS,
        strict=True,
    )
)
_RESPONSIBILITY_FIELDS: Final = dict(
    zip(
        (
            "service_type",
            "asset_owner",
            "asset_department",
            "port_owner",
            "department",
        ),
        WARNING_HEADERS,
        strict=True,
    )
)
_OPTIONAL_HEADERS: Final = set(OPTIONAL_HEADERS)
_KNOWN_HEADERS: Final = (
    set(_REQUIRED_FIELDS.values())
    | set(_RESPONSIBILITY_FIELDS.values())
    | _OPTIONAL_HEADERS
)
_DECIMAL_PORT = re.compile(r"[0-9]+", flags=re.ASCII)


@dataclass(frozen=True)
class CustomerUploadWarning:
    code: str
    field: str | None
    count: int


@dataclass(frozen=True)
class CustomerUploadValidationResult:
    record_count: int
    warnings: tuple[CustomerUploadWarning, ...]


class CustomerUploadValidationError(Exception):
    """A stable rejection that never carries workbook or parser content."""

    def __init__(
        self, code: str, *, field: str | None = None, row: int | None = None
    ) -> None:
        super().__init__(code)
        self.code = code
        self.field = field
        self.row = row


def _reject(
    code: str, *, field: str | None = None, row: int | None = None
) -> CustomerUploadValidationError:
    return CustomerUploadValidationError(code, field=field, row=row)


def _is_blank(value: object) -> bool:
    return value is None or isinstance(value, str) and not value.strip()


def _is_empty_url(value: object) -> bool:
    return value is None or value == ""


def _row_value(row: Sequence[Any], index: int) -> object:
    return row[index] if index < len(row) else None


def _validate_ip(value: object, row_number: int) -> None:
    if not isinstance(value, str):
        raise _reject("invalid_required_value", field="asset_ip", row=row_number)
    candidate = value.strip()
    valid_ip = "%" not in candidate
    try:
        if valid_ip:
            ipaddress.ip_address(candidate)
    except ValueError:
        valid_ip = False
    if not valid_ip:
        raise _reject("invalid_required_value", field="asset_ip", row=row_number)


def _parse_port(value: object, field: str, row_number: int) -> int:
    if isinstance(value, bool):
        raise _reject("invalid_required_value", field=field, row=row_number)
    if isinstance(value, int):
        port = value
    elif isinstance(value, str) and _DECIMAL_PORT.fullmatch(value.strip()):
        normalized = value.strip().lstrip("0") or "0"
        if len(normalized) > 5 or (
            len(normalized) == 5 and normalized > "65535"
        ):
            raise _reject("invalid_required_value", field=field, row=row_number)
        port = int(normalized)
    else:
        raise _reject("invalid_required_value", field=field, row=row_number)
    if not 1 <= port <= 65_535:
        raise _reject("invalid_required_value", field=field, row=row_number)
    return port


def _validate_required_row(
    row: Sequence[Any], field_indexes: dict[str, int], row_number: int
) -> None:
    _validate_ip(_row_value(row, field_indexes["asset_ip"]), row_number)
    start_port = _parse_port(
        _row_value(row, field_indexes["start_port"]), "start_port", row_number
    )
    end_port = _parse_port(
        _row_value(row, field_indexes["end_port"]), "end_port", row_number
    )
    if start_port != end_port:
        raise _reject(
            "invalid_required_value", field="end_port", row=row_number
        )

    web_value = _row_value(row, field_indexes["is_web"])
    if not isinstance(web_value, str) or web_value.strip() not in {"是", "否", "无"}:
        raise _reject("invalid_required_value", field="is_web", row=row_number)
    url_value = _row_value(row, field_indexes["web_url"])
    if web_value.strip() == "是":
        if not isinstance(url_value, str) or not url_value.strip():
            raise _reject(
                "invalid_required_value", field="web_url", row=row_number
            )
    elif not _is_empty_url(url_value):
        raise _reject("invalid_required_value", field="web_url", row=row_number)


def _parse_rows(
    worksheet: Any, bounds: PreflightResult
) -> CustomerUploadValidationResult:
    if bounds.max_column > bounds.max_header_column:
        raise _reject("missing_required_structure", row=1)
    rows = worksheet.iter_rows(
        min_row=1,
        max_row=max(1, bounds.max_row),
        min_col=1,
        max_col=max(1, bounds.max_column, bounds.max_header_column),
        values_only=True,
    )
    header_row: Sequence[Any] | None
    try:
        header_row = next(rows)
    except StopIteration:
        header_row = None
    if header_row is None:
        raise _reject("missing_required_structure")

    header_values = list(header_row)

    headers: list[str] = []
    for value in header_values:
        if not isinstance(value, str) or not value.strip():
            raise _reject("missing_required_structure", row=1)
        headers.append(value)
    if len(headers) != len(set(headers)):
        raise _reject("missing_required_structure", row=1)

    header_indexes = {header: index for index, header in enumerate(headers)}
    if not any(header in header_indexes for header in _REQUIRED_FIELDS.values()):
        raise _reject("missing_required_structure", row=1)

    field_indexes: dict[str, int] = {}
    for field, header in _REQUIRED_FIELDS.items():
        if header not in header_indexes:
            raise _reject("missing_required_structure", field=field, row=1)
        field_indexes[field] = header_indexes[header]

    responsibility_indexes = {
        field: header_indexes.get(header)
        for field, header in _RESPONSIBILITY_FIELDS.items()
    }
    warning_counts = dict.fromkeys(_RESPONSIBILITY_FIELDS, 0)
    record_count = 0
    for row_number, row in enumerate(rows, start=2):
        if all(value is None for value in row):
            continue
        if any(value is not None for value in row[len(headers) :]):
            raise _reject("missing_required_structure", row=1)
        _validate_required_row(row, field_indexes, row_number)
        record_count += 1
        for field, index in responsibility_indexes.items():
            if index is None or _is_blank(_row_value(row, index)):
                warning_counts[field] += 1

    if record_count == 0:
        raise _reject("missing_required_structure")

    warnings = [
        CustomerUploadWarning(
            code="missing_responsibility_value", field=field, count=count
        )
        for field, count in warning_counts.items()
        if count
    ]
    extra_column_count = sum(header not in _KNOWN_HEADERS for header in headers)
    if extra_column_count:
        warnings.append(
            CustomerUploadWarning(
                code="extra_columns_ignored", field=None, count=extra_column_count
            )
        )
    return CustomerUploadValidationResult(
        record_count=record_count, warnings=tuple(warnings)
    )


def _check_parser_contract() -> None:
    if (
        openpyxl.__version__ != "3.1.5"
        or defusedxml.__version__ != "0.7.1"
        or OPENPYXL_DEFUSEDXML is not True
    ):
        raise RuntimeError("XLSX parser runtime contract is not satisfied")


def validate_customer_upload_workbook(
    path: Path,
) -> CustomerUploadValidationResult:
    """Validate one local XLSX path against the fixed default CustomerUpload v1."""
    _check_parser_contract()
    stat_failed = False
    try:
        file_size = path.stat().st_size
    except OSError:
        stat_failed = True
        file_size = 0
    if stat_failed:
        raise _reject("malformed_workbook")
    if file_size > MAX_WORKBOOK_BYTES:
        raise _reject("upload_too_large")

    preflight_error: PreflightError | None = None
    preflight_result: PreflightResult | None = None
    try:
        preflight_result = preflight_xlsx(path)
    except PreflightError as error:
        preflight_error = error
    if preflight_error is not None:
        raise _reject(preflight_error.code, row=preflight_error.row)
    if preflight_result is None:
        raise _reject("malformed_workbook")

    workbook: Any | None = None
    result: CustomerUploadValidationResult | None = None
    parser_failed = False
    try:
        workbook = load_workbook(
            path, read_only=True, data_only=False, keep_links=True
        )
        if len(workbook.worksheets) != 1:
            raise _reject("unsupported_workbook_feature")
        worksheet = workbook.worksheets[0]
        if worksheet.sheet_state != "visible":
            raise _reject("unsupported_workbook_feature")
        result = _parse_rows(worksheet, preflight_result)
    except CustomerUploadValidationError:
        raise
    except Exception:
        parser_failed = True
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                parser_failed = True
    if parser_failed or result is None:
        raise _reject("malformed_workbook")
    return result
