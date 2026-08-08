"""Deterministic IP facts derived from two complete source artifacts.

This module is deliberately a pure processing seam.  It reads the already
accepted CustomerUpload workbook and the already complete CloudAtlas snapshot,
but it does not create database rows, findings, evidence, or candidate state.
"""

from __future__ import annotations

import hashlib
import ipaddress
import json
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Final, TypeGuard, cast

from openpyxl import load_workbook  # type: ignore[import-untyped]

from app.domain.customer_upload_profiles import REQUIRED_HEADERS
from app.domain.customer_upload_validator import (
    CustomerUploadValidationError,
    validate_customer_upload_workbook,
)

CUSTOMER_UPLOAD_SOURCE_TYPE: Final = "CUSTOMER_UPLOAD"
CLOUDATLAS_SOURCE_TYPE: Final = "CLOUDATLAS"
IP_RESOURCE_TYPE: Final = "IP"
IP_PROCESSING_CONTRACT_VERSION: Final = "ip-v1"
CLOUDATLAS_SNAPSHOT_SCHEMA: Final = (
    "exposure-agent.cloudatlas-ip-assets.snapshot.v1"
)

_CUSTOMER_IP_HEADER: Final = REQUIRED_HEADERS[0]
_CLOUDATLAS_PAGE_KEYS: Final = frozenset({"items", "page", "size", "total"})
_CLOUDATLAS_ITEM_KEYS: Final = frozenset({"id", "ip", "status"})

type CloudAtlasArtifactInput = Path | str | bytes | bytearray | Mapping[str, Any]


class IPRecordContractError(Exception):
    """A stable source-record contract error with no source content in its text."""

    def __init__(
        self,
        code: str,
        *,
        source_type: str | None = None,
        source_record_key: str | None = None,
        field: str | None = None,
        row: int | None = None,
    ) -> None:
        super().__init__(code)
        self.code = code
        self.source_type = source_type
        self.source_record_key = source_record_key
        self.field = field
        self.row = row


@dataclass(frozen=True, slots=True)
class IPObservation:
    """One source record, without any cross-source deduplication."""

    source_type: str
    source_record_key: str
    raw_ip: str
    canonical_ip: str
    cloudatlas_asset_id: str | None = None
    cloudatlas_status: str | None = None

    def as_dict(self) -> dict[str, str]:
        value = {
            "source_type": self.source_type,
            "source_record_key": self.source_record_key,
            "raw_ip": self.raw_ip,
            "canonical_ip": self.canonical_ip,
        }
        if self.source_type == CLOUDATLAS_SOURCE_TYPE:
            # CloudAtlas id/status are source metadata confirmed by its
            # protobuf contract; customer rows have no corresponding fields.
            value["cloudatlas_asset_id"] = self.cloudatlas_asset_id or ""
            value["cloudatlas_status"] = self.cloudatlas_status or ""
        return value


@dataclass(frozen=True, slots=True)
class IPResource:
    """A Project-scoped IP identity represented by its canonical key."""

    canonical_key: str

    def as_dict(self) -> dict[str, str]:
        return {
            "resource_type": IP_RESOURCE_TYPE,
            "canonical_key": self.canonical_key,
        }


@dataclass(frozen=True, slots=True)
class IPObservationResourceLink:
    """The exact resolution of one ordered observation to one IP resource."""

    observation_index: int
    source_type: str
    source_record_key: str
    resource_key: str
    processing_contract_version: str

    def as_dict(self) -> dict[str, int | str]:
        return {
            "observation_index": self.observation_index,
            "source_type": self.source_type,
            "source_record_key": self.source_record_key,
            "resource_key": self.resource_key,
            "processing_contract_version": self.processing_contract_version,
        }


@dataclass(frozen=True, slots=True)
class IPResolution:
    """Stable Project resources and exact links for an ordered observation list."""

    resources: tuple[IPResource, ...]
    links: tuple[IPObservationResourceLink, ...]

    def as_dict(self) -> dict[str, list[dict[str, int | str]] | list[dict[str, str]]]:
        return {
            "resources": [resource.as_dict() for resource in self.resources],
            "links": [link.as_dict() for link in self.links],
        }


@dataclass(frozen=True, slots=True)
class IPDifferenceSet:
    """Set conclusions; duplicate source observations do not duplicate them."""

    cloudatlas_only: tuple[str, ...]
    customer_upload_only: tuple[str, ...]
    matched: tuple[str, ...]

    def as_dict(self) -> dict[str, list[str]]:
        return {
            "cloudatlas_only": list(self.cloudatlas_only),
            "customer_upload_only": list(self.customer_upload_only),
            "matched": list(self.matched),
        }


@dataclass(frozen=True, slots=True)
class IPConsistencyResult:
    """All deterministic facts produced by the processing seam."""

    processing_contract_version: str
    observations: tuple[IPObservation, ...]
    resources: tuple[IPResource, ...]
    links: tuple[IPObservationResourceLink, ...]
    differences: IPDifferenceSet
    output_hash: str

    def as_dict(self) -> dict[str, Any]:
        return _result_payload(
            self.processing_contract_version,
            self.observations,
            self.resources,
            self.links,
            self.differences,
        )


def _error(
    code: str,
    *,
    source_type: str | None = None,
    source_record_key: str | None = None,
    field: str | None = None,
    row: int | None = None,
) -> IPRecordContractError:
    return IPRecordContractError(
        code,
        source_type=source_type,
        source_record_key=source_record_key,
        field=field,
        row=row,
    )


def _is_int(value: object) -> TypeGuard[int]:
    return isinstance(value, int) and not isinstance(value, bool)


def _canonical_ip(
    value: object,
    *,
    source_type: str,
    source_record_key: str,
) -> str:
    if not isinstance(value, str):
        raise _error(
            "invalid_ip",
            source_type=source_type,
            source_record_key=source_record_key,
            field="ip",
        )
    candidate = value.strip()
    # ipaddress accepts scoped IPv6 addresses (for example %eth0), but a
    # source IP literal in this contract has no interface scope.
    if not candidate or "%" in candidate:
        raise _error(
            "invalid_ip",
            source_type=source_type,
            source_record_key=source_record_key,
            field="ip",
        )
    try:
        address = ipaddress.ip_address(candidate)
    except ValueError:
        raise _error(
            "invalid_ip",
            source_type=source_type,
            source_record_key=source_record_key,
            field="ip",
        ) from None
    if isinstance(address, ipaddress.IPv4Address):
        return str(address)
    if address.ipv4_mapped is not None:
        # ipv4_mapped is non-None only for the standard ::ffff:0:0/96
        # mapped range.  No other IPv6 address is folded.
        return str(address.ipv4_mapped)
    return address.compressed


def _ip_sort_key(value: str) -> tuple[int, int]:
    address = ipaddress.ip_address(value)
    return address.version, int(address)


def _validate_processing_contract_version(value: str) -> str:
    if not isinstance(value, str):
        raise _error("processing_contract_version_invalid")
    normalized = value.strip()
    if not normalized or len(normalized) > 100:
        raise _error("processing_contract_version_invalid")
    return normalized


def _customer_observations(path_value: Path | str) -> tuple[IPObservation, ...]:
    try:
        path = Path(path_value)
    except TypeError:
        raise _error("customer_snapshot_contract_invalid", source_type=CUSTOMER_UPLOAD_SOURCE_TYPE) from None

    try:
        validation = validate_customer_upload_workbook(path)
    except CustomerUploadValidationError as error:
        source_record_key = f"row:{error.row}" if error.row is not None else None
        raise _error(
            error.code,
            source_type=CUSTOMER_UPLOAD_SOURCE_TYPE,
            source_record_key=source_record_key,
            field=error.field,
            row=error.row,
        ) from None
    except Exception:
        raise _error(
            "customer_snapshot_contract_invalid",
            source_type=CUSTOMER_UPLOAD_SOURCE_TYPE,
        ) from None

    workbook: Any | None = None
    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_links=True,
        )
        if len(workbook.worksheets) != 1:
            raise _error(
                "customer_snapshot_contract_invalid",
                source_type=CUSTOMER_UPLOAD_SOURCE_TYPE,
            )
        worksheet = workbook.worksheets[0]
        if worksheet.sheet_state != "visible":
            raise _error(
                "customer_snapshot_contract_invalid",
                source_type=CUSTOMER_UPLOAD_SOURCE_TYPE,
            )
        max_row = max(1, worksheet.max_row or 1)
        max_column = max(1, worksheet.max_column or 1)
        rows = worksheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_column,
            values_only=True,
        )
        try:
            header_row = list(next(rows))
        except StopIteration:
            raise _error(
                "customer_snapshot_contract_invalid",
                source_type=CUSTOMER_UPLOAD_SOURCE_TYPE,
            ) from None
        try:
            ip_index = header_row.index(_CUSTOMER_IP_HEADER)
        except ValueError:
            raise _error(
                "customer_snapshot_contract_invalid",
                source_type=CUSTOMER_UPLOAD_SOURCE_TYPE,
            ) from None

        observations: list[IPObservation] = []
        for row_number, row in enumerate(rows, start=2):
            if all(value is None for value in row):
                continue
            raw_ip = row[ip_index] if ip_index < len(row) else None
            source_record_key = f"row:{row_number}"
            canonical_ip = _canonical_ip(
                raw_ip,
                source_type=CUSTOMER_UPLOAD_SOURCE_TYPE,
                source_record_key=source_record_key,
            )
            observations.append(
                IPObservation(
                    source_type=CUSTOMER_UPLOAD_SOURCE_TYPE,
                    source_record_key=source_record_key,
                    raw_ip=cast(str, raw_ip),
                    canonical_ip=canonical_ip,
                )
            )
        if len(observations) != validation.record_count:
            raise _error(
                "customer_snapshot_record_count_mismatch",
                source_type=CUSTOMER_UPLOAD_SOURCE_TYPE,
            )
        return tuple(observations)
    except IPRecordContractError:
        raise
    except Exception:
        raise _error(
            "customer_snapshot_contract_invalid",
            source_type=CUSTOMER_UPLOAD_SOURCE_TYPE,
        ) from None
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                # The validator already completed the contract check.  A
                # close failure must not leak parser details to the caller.
                pass


def _load_cloudatlas_payload(source: CloudAtlasArtifactInput) -> Mapping[str, Any]:
    if isinstance(source, Mapping):
        mapping_source: object = source
        return dict(cast(Mapping[str, Any], mapping_source))
    if isinstance(source, (bytes, bytearray)):
        raw = bytes(source)
    else:
        try:
            raw = Path(source).read_bytes()
        except (OSError, TypeError):
            raise _error(
                "cloudatlas_snapshot_contract_invalid",
                source_type=CLOUDATLAS_SOURCE_TYPE,
            ) from None
    try:
        payload = json.loads(raw.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError):
        raise _error(
            "cloudatlas_snapshot_contract_invalid",
            source_type=CLOUDATLAS_SOURCE_TYPE,
        ) from None
    if not isinstance(payload, Mapping):
        raise _error(
            "cloudatlas_snapshot_contract_invalid",
            source_type=CLOUDATLAS_SOURCE_TYPE,
        )
    return payload


def _cloudatlas_observations(
    source: CloudAtlasArtifactInput,
) -> tuple[IPObservation, ...]:
    payload = _load_cloudatlas_payload(source)
    if set(payload) != {"pages", "schema"}:
        raise _error(
            "cloudatlas_snapshot_contract_invalid",
            source_type=CLOUDATLAS_SOURCE_TYPE,
        )
    if payload.get("schema") != CLOUDATLAS_SNAPSHOT_SCHEMA:
        raise _error(
            "cloudatlas_snapshot_schema_invalid",
            source_type=CLOUDATLAS_SOURCE_TYPE,
        )
    pages = payload.get("pages")
    if not isinstance(pages, list) or not pages:
        raise _error(
            "snapshot_incomplete",
            source_type=CLOUDATLAS_SOURCE_TYPE,
        )

    expected_size: int | None = None
    expected_total: int | None = None
    observations: list[IPObservation] = []
    for expected_page_number, page_value in enumerate(pages, start=1):
        if not isinstance(page_value, Mapping) or set(page_value) != _CLOUDATLAS_PAGE_KEYS:
            raise _error(
                "cloudatlas_snapshot_contract_invalid",
                source_type=CLOUDATLAS_SOURCE_TYPE,
            )
        page = page_value.get("page")
        size = page_value.get("size")
        total = page_value.get("total")
        items = page_value.get("items")
        if not _is_int(page) or not _is_int(size) or not _is_int(total):
            raise _error(
                "cloudatlas_snapshot_contract_invalid",
                source_type=CLOUDATLAS_SOURCE_TYPE,
            )
        if (
            page != expected_page_number
            or size <= 0
            or total < 0
            or not isinstance(items, list)
        ):
            raise _error(
                "snapshot_incomplete",
                source_type=CLOUDATLAS_SOURCE_TYPE,
            )
        if expected_size is None:
            expected_size = size
            expected_total = total
        assert expected_size is not None
        assert expected_total is not None
        if size != expected_size or total != expected_total:
            raise _error(
                "snapshot_incomplete",
                source_type=CLOUDATLAS_SOURCE_TYPE,
            )
        if len(observations) == expected_total and expected_page_number > 1:
            raise _error(
                "snapshot_incomplete",
                source_type=CLOUDATLAS_SOURCE_TYPE,
            )
        if len(items) > size:
            raise _error(
                "snapshot_incomplete",
                source_type=CLOUDATLAS_SOURCE_TYPE,
            )
        if expected_total == 0 and items:
            raise _error(
                "snapshot_incomplete",
                source_type=CLOUDATLAS_SOURCE_TYPE,
            )
        if len(observations) + len(items) > expected_total:
            raise _error(
                "snapshot_incomplete",
                source_type=CLOUDATLAS_SOURCE_TYPE,
            )
        if expected_total and not items and len(observations) < expected_total:
            raise _error(
                "snapshot_incomplete",
                source_type=CLOUDATLAS_SOURCE_TYPE,
            )

        for item_index, item_value in enumerate(items):
            source_record_key = f"page:{page}:item:{item_index}"
            if not isinstance(item_value, Mapping) or set(item_value) != _CLOUDATLAS_ITEM_KEYS:
                raise _error(
                    "invalid_item_contract",
                    source_type=CLOUDATLAS_SOURCE_TYPE,
                    source_record_key=source_record_key,
                )
            asset_id = item_value.get("id")
            raw_ip = item_value.get("ip")
            status = item_value.get("status")
            if (
                not isinstance(asset_id, str)
                or not asset_id.strip()
                or not isinstance(raw_ip, str)
                or not isinstance(status, str)
            ):
                raise _error(
                    "invalid_item_contract",
                    source_type=CLOUDATLAS_SOURCE_TYPE,
                    source_record_key=source_record_key,
                )
            if status != "valid":
                raise _error(
                    "invalid_status",
                    source_type=CLOUDATLAS_SOURCE_TYPE,
                    source_record_key=source_record_key,
                    field="status",
                )
            canonical_ip = _canonical_ip(
                raw_ip,
                source_type=CLOUDATLAS_SOURCE_TYPE,
                source_record_key=source_record_key,
            )
            observations.append(
                IPObservation(
                    source_type=CLOUDATLAS_SOURCE_TYPE,
                    source_record_key=source_record_key,
                    raw_ip=raw_ip,
                    canonical_ip=canonical_ip,
                    cloudatlas_asset_id=asset_id,
                    cloudatlas_status=status,
                )
            )

    if expected_total is None or len(observations) != expected_total:
        raise _error(
            "snapshot_incomplete",
            source_type=CLOUDATLAS_SOURCE_TYPE,
        )
    if expected_total == 0 and len(pages) != 1:
        raise _error(
            "snapshot_incomplete",
            source_type=CLOUDATLAS_SOURCE_TYPE,
        )
    return tuple(observations)


def normalize_ip(value: str) -> str:
    """Return the contract Canonical IP for a standalone IP literal."""

    return _canonical_ip(
        value,
        source_type="IP",
        source_record_key="value",
    )


def resolve_ip_observations(
    observations: Sequence[IPObservation],
    *,
    processing_contract_version: str = IP_PROCESSING_CONTRACT_VERSION,
) -> IPResolution:
    """Resolve every observation exactly once without removing duplicates."""

    version = _validate_processing_contract_version(processing_contract_version)
    canonical_keys = {observation.canonical_ip for observation in observations}
    resources = tuple(
        IPResource(canonical_key=canonical_ip)
        for canonical_ip in sorted(canonical_keys, key=_ip_sort_key)
    )
    links = tuple(
        IPObservationResourceLink(
            observation_index=index,
            source_type=observation.source_type,
            source_record_key=observation.source_record_key,
            resource_key=observation.canonical_ip,
            processing_contract_version=version,
        )
        for index, observation in enumerate(observations)
    )
    return IPResolution(resources=resources, links=links)


def check_ip_differences(
    observations: Sequence[IPObservation],
) -> IPDifferenceSet:
    """Compute set conclusions from source observations in linear time."""

    customer_ips: set[str] = set()
    cloudatlas_ips: set[str] = set()
    for observation in observations:
        if observation.source_type == CUSTOMER_UPLOAD_SOURCE_TYPE:
            customer_ips.add(observation.canonical_ip)
        elif observation.source_type == CLOUDATLAS_SOURCE_TYPE:
            cloudatlas_ips.add(observation.canonical_ip)
        else:
            raise _error(
                "observation_source_type_invalid",
                source_type=observation.source_type,
                source_record_key=observation.source_record_key,
            )
    return IPDifferenceSet(
        cloudatlas_only=tuple(
            sorted(cloudatlas_ips - customer_ips, key=_ip_sort_key)
        ),
        customer_upload_only=tuple(
            sorted(customer_ips - cloudatlas_ips, key=_ip_sort_key)
        ),
        matched=tuple(sorted(customer_ips & cloudatlas_ips, key=_ip_sort_key)),
    )


def _result_payload(
    processing_contract_version: str,
    observations: Sequence[IPObservation],
    resources: Sequence[IPResource],
    links: Sequence[IPObservationResourceLink],
    differences: IPDifferenceSet,
) -> dict[str, Any]:
    return {
        "processing_contract_version": processing_contract_version,
        "observations": [observation.as_dict() for observation in observations],
        "resources": [resource.as_dict() for resource in resources],
        "links": [link.as_dict() for link in links],
        "differences": differences.as_dict(),
    }


def _output_hash(payload: Mapping[str, Any]) -> str:
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        separators=(",", ":"),
        sort_keys=True,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def process_ip_snapshots(
    customer_upload_artifact: Path | str,
    cloudatlas_snapshot_artifact: CloudAtlasArtifactInput,
    *,
    processing_contract_version: str = IP_PROCESSING_CONTRACT_VERSION,
) -> IPConsistencyResult:
    """Process two complete immutable artifacts atomically in memory.

    The function builds no externally visible state until both source records
    and every record-level contract have passed.  Any ``IPRecordContractError``
    therefore represents a failed processing attempt with no partial result.
    """

    version = _validate_processing_contract_version(processing_contract_version)
    customer_observations = _customer_observations(customer_upload_artifact)
    cloudatlas_observations = _cloudatlas_observations(cloudatlas_snapshot_artifact)
    observations = customer_observations + cloudatlas_observations
    resolution = resolve_ip_observations(
        observations,
        processing_contract_version=version,
    )
    differences = check_ip_differences(observations)
    payload = _result_payload(
        version,
        observations,
        resolution.resources,
        resolution.links,
        differences,
    )
    return IPConsistencyResult(
        processing_contract_version=version,
        observations=observations,
        resources=resolution.resources,
        links=resolution.links,
        differences=differences,
        output_hash=_output_hash(payload),
    )
