import re
import sys
from collections.abc import Callable, Sequence
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet._read_only import (  # type: ignore[import-untyped]
    ReadOnlyWorksheet,
)

from app.domain.customer_upload_validator import (
    MAX_WORKBOOK_BYTES,
    CustomerUploadValidationError,
    CustomerUploadWarning,
    validate_customer_upload_workbook,
)

_REPOSITORY_ROOT = Path(__file__).parents[3]
sys.path.insert(0, str(_REPOSITORY_ROOT))
from investigations.issue_33.probe import build_fixture  # type: ignore[import-not-found]  # noqa: E402, I001


DEFAULT_HEADERS = [
    "资产IP",
    "起始端口",
    "结束端口",
    "是否web界面",
    "web界面url",
    "服务类型",
    "资产负责人",
    "资产所属部门",
    "端口负责人",
    "部门",
]


def _save_workbook(path: Path, rows: Sequence[Sequence[object]]) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def _rewrite_zip_member(
    path: Path, name: str, transform: Callable[[bytes], bytes]
) -> None:
    replacement = path.with_suffix(".rewritten.xlsx")
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED) as target,
    ):
        for info in source.infolist():
            data = source.read(info)
            if info.filename == name:
                data = transform(data)
            target.writestr(info, data)
    replacement.replace(path)


def _move_formula_to_responsibility_field(
    data: bytes, encoding: str
) -> bytes:
    data = re.sub(rb'<c r="K2".*?</c>', b"", data, count=1)
    data = re.sub(
        rb'<c r="F2".*?</c>',
        b'<c r="F2"><f>1+1</f><v /></c>',
        data,
        count=1,
    )
    return data.decode("utf-8").encode(encoding)


def test_valid_default_v1_workbook_returns_record_count(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "input.xlsx",
        [
            DEFAULT_HEADERS,
            [
                "192.0.2.10",
                443,
                443,
                "是",
                "fixture.example.invalid",
                "Fixture Service",
                "Example Owner",
                "Example Department",
                "Example Port Owner",
                "Example Operations",
            ],
            [
                "2001:db8::10",
                "22",
                "22",
                "否",
                None,
                "Fixture SSH",
                "Example Owner",
                "Example Department",
                "Example Port Owner",
                "Example Operations",
            ],
        ],
    )

    result = validate_customer_upload_workbook(path)

    assert result.record_count == 2
    assert result.warnings == ()


def _assert_rejected(
    path: Path,
    code: str,
    *,
    field: str | None = None,
    row: int | None = None,
) -> None:
    with pytest.raises(CustomerUploadValidationError) as caught:
        validate_customer_upload_workbook(path)

    assert caught.value.code == code
    assert caught.value.field == field
    assert caught.value.row == row
    assert str(caught.value) == code


@pytest.mark.parametrize(
    ("headers", "field"),
    [
        (DEFAULT_HEADERS[1:], "asset_ip"),
        (["资产IP ", *DEFAULT_HEADERS[1:]], "asset_ip"),
        ([*DEFAULT_HEADERS, "资产IP"], None),
        ([*DEFAULT_HEADERS[:2], None, *DEFAULT_HEADERS[3:]], None),
        ([*DEFAULT_HEADERS, "   "], None),
    ],
)
def test_headers_must_match_default_v1_exactly_and_be_unique(
    tmp_path: Path, headers: list[object], field: str | None
) -> None:
    path = _save_workbook(tmp_path / "headers.xlsx", [headers, ["partial"]])

    _assert_rejected(
        path, "missing_required_structure", field=field, row=1
    )


def test_workbook_requires_one_visible_sheet(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "sheets.xlsx",
        [DEFAULT_HEADERS, ["192.0.2.1", 443, 443, "否", None]],
    )
    workbook = load_workbook(path)
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    workbook.save(path)
    workbook.close()

    _assert_rejected(path, "unsupported_workbook_feature")


@pytest.mark.parametrize("merged_range", ["A1:B1", "A1:A2"])
def test_merged_header_is_rejected(
    tmp_path: Path, merged_range: str
) -> None:
    path = _save_workbook(
        tmp_path / "merged.xlsx",
        [DEFAULT_HEADERS, ["192.0.2.1", 443, 443, "否", None]],
    )
    workbook = load_workbook(path)
    workbook.active.merge_cells(merged_range)
    workbook.save(path)
    workbook.close()

    _assert_rejected(path, "missing_required_structure", row=1)


def test_title_row_before_headers_is_rejected_as_multi_row_header(
    tmp_path: Path,
) -> None:
    path = _save_workbook(
        tmp_path / "title.xlsx",
        [["Asset Inventory"], DEFAULT_HEADERS, ["192.0.2.1", 443, 443, "否", None]],
    )

    _assert_rejected(path, "missing_required_structure", row=1)


@pytest.mark.parametrize(
    ("value", "field"),
    [
        ("192.0.2.1/32", "asset_ip"),
        ("192.0.2.1-192.0.2.2", "asset_ip"),
        ("fixture.example.invalid", "asset_ip"),
        ("192.0.2.1,198.51.100.1", "asset_ip"),
        (0, "start_port"),
        (65_536, "start_port"),
        (1.5, "start_port"),
        ("1.0", "start_port"),
        ("0x50", "start_port"),
        ("1-2", "start_port"),
    ],
)
def test_invalid_ip_and_port_values_are_rejected_with_canonical_field(
    tmp_path: Path, value: object, field: str
) -> None:
    row: list[object] = ["192.0.2.1", 443, 443, "否", None]
    row[0 if field == "asset_ip" else 1] = value
    path = _save_workbook(tmp_path / "value.xlsx", [DEFAULT_HEADERS, row])

    _assert_rejected(path, "invalid_required_value", field=field, row=2)


def test_scoped_ipv6_is_rejected_as_non_literal(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "scoped-ipv6.xlsx",
        [DEFAULT_HEADERS, ["fe80::1%eth0", 443, 443, "否", None]],
    )

    _assert_rejected(
        path, "invalid_required_value", field="asset_ip", row=2
    )


def test_oversized_decimal_port_has_canonical_field_and_row(
    tmp_path: Path,
) -> None:
    path = _save_workbook(
        tmp_path / "oversized-port.xlsx",
        [DEFAULT_HEADERS, ["192.0.2.1", "9" * 5_000, 443, "否", None]],
    )

    _assert_rejected(
        path, "invalid_required_value", field="start_port", row=2
    )


def test_port_range_is_rejected(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "range.xlsx",
        [DEFAULT_HEADERS, ["192.0.2.1", 80, 443, "否", None]],
    )

    _assert_rejected(
        path, "invalid_required_value", field="end_port", row=2
    )


@pytest.mark.parametrize(
    ("web_flag", "url", "field"),
    [
        ("true", None, "is_web"),
        ("是", None, "web_url"),
        ("否", "fixture.example.invalid", "web_url"),
        ("无", "fixture.example.invalid", "web_url"),
    ],
)
def test_web_flag_and_url_combination_is_strict(
    tmp_path: Path, web_flag: str, url: str | None, field: str
) -> None:
    path = _save_workbook(
        tmp_path / "web.xlsx",
        [DEFAULT_HEADERS, ["192.0.2.1", 443, 443, web_flag, url]],
    )

    _assert_rejected(path, "invalid_required_value", field=field, row=2)


def test_blank_rows_are_ignored_but_partial_rows_fail_at_actual_xlsx_row(
    tmp_path: Path,
) -> None:
    path = _save_workbook(
        tmp_path / "partial.xlsx",
        [
            DEFAULT_HEADERS,
            ["192.0.2.1", 443, 443, "否", None],
            [None] * len(DEFAULT_HEADERS),
            [None, None, None, None, None, "partial responsibility"],
        ],
    )

    _assert_rejected(
        path, "invalid_required_value", field="asset_ip", row=4
    )


def test_whitespace_only_row_is_partial_not_physically_blank(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "whitespace-row.xlsx",
        [
            DEFAULT_HEADERS,
            ["192.0.2.1", 443, 443, "否", None],
            ["   "] + [None] * (len(DEFAULT_HEADERS) - 1),
        ],
    )

    _assert_rejected(
        path, "invalid_required_value", field="asset_ip", row=3
    )


def test_multiline_undefined_header_is_counted_as_an_extra_column(
    tmp_path: Path,
) -> None:
    path = _save_workbook(
        tmp_path / "extra.xlsx",
        [
            [*DEFAULT_HEADERS, "Extra\nDescription"],
            [
                "192.0.2.1",
                443,
                443,
                "否",
                None,
                "Service",
                "Owner",
                "Department",
                "Port Owner",
                "Operations",
                "Fixture",
            ],
        ],
    )

    result = validate_customer_upload_workbook(path)

    assert result.warnings == (
        CustomerUploadWarning("extra_columns_ignored", None, 1),
    )


@pytest.mark.parametrize("coordinate", ["Z2", "XFD2000"])
def test_style_only_cells_outside_data_columns_are_ignored(
    tmp_path: Path, coordinate: str
) -> None:
    path = _save_workbook(
        tmp_path / "styled.xlsx",
        [DEFAULT_HEADERS, ["192.0.2.1", 443, 443, "否", None]],
    )
    workbook = load_workbook(path)
    workbook.active[coordinate].number_format = "0.00"
    workbook.save(path)
    workbook.close()

    result = validate_customer_upload_workbook(path)

    assert result.record_count == 1


def test_warning_counts_cover_missing_columns_empty_values_and_extra_columns(
    tmp_path: Path,
) -> None:
    headers = [
        *DEFAULT_HEADERS[:6],
        "资产负责人",
        "序号",
        "Unmapped Fixture Column",
    ]
    path = _save_workbook(
        tmp_path / "warnings.xlsx",
        [
            headers,
            ["192.0.2.1", 443, 443, "否", None, None, "Owner", None, "x"],
            ["2001:db8::1", "22", "22", "无", None, "SSH", "  ", None, "y"],
            [None] * len(headers),
        ],
    )

    result = validate_customer_upload_workbook(path)

    assert result.record_count == 2
    assert result.warnings == (
        CustomerUploadWarning("missing_responsibility_value", "service_type", 1),
        CustomerUploadWarning("missing_responsibility_value", "asset_owner", 1),
        CustomerUploadWarning("missing_responsibility_value", "asset_department", 2),
        CustomerUploadWarning("missing_responsibility_value", "port_owner", 2),
        CustomerUploadWarning("missing_responsibility_value", "department", 2),
        CustomerUploadWarning("extra_columns_ignored", None, 1),
    )


def test_workbook_requires_a_non_empty_data_row(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "empty.xlsx", [DEFAULT_HEADERS, [None] * len(DEFAULT_HEADERS)]
    )

    _assert_rejected(path, "missing_required_structure")


@pytest.mark.parametrize(
    ("fixture_name", "expected_records"),
    [
        ("default_v1", 3),
        ("near_request_limit", 3),
        ("row_shared_style_boundary", 50_000),
    ],
)
def test_issue_33_normal_boundaries_are_accepted(
    tmp_path: Path, fixture_name: str, expected_records: int
) -> None:
    path = build_fixture(fixture_name, tmp_path / "boundary.xlsx")

    result = validate_customer_upload_workbook(path)

    assert result.record_count == expected_records
    assert result.warnings == ()


@pytest.mark.parametrize(
    "fixture_name",
    ["entry_count_bomb", "compression_bomb", "total_size_bomb"],
)
def test_issue_33_zip_bombs_have_one_stable_rejection(
    tmp_path: Path, fixture_name: str
) -> None:
    path = build_fixture(fixture_name, tmp_path / "bomb.xlsx")

    _assert_rejected(path, "workbook_resource_limit")


@pytest.mark.parametrize(
    "fixture_name",
    [
        "formula",
        "relocated_formula_without_content_type",
        "padded_relocated_formula",
        "table_formula",
        "conditional_format_formula",
        "data_validation_formula",
        "external_link",
        "data_connection",
        "hidden_sheet",
        "hidden_chartsheet",
        "visible_chartsheet",
        "embedded_object",
        "vml_button",
    ],
)
def test_issue_33_active_content_has_one_stable_rejection(
    tmp_path: Path, fixture_name: str
) -> None:
    path = build_fixture(fixture_name, tmp_path / "active.xlsx")

    _assert_rejected(path, "unsupported_workbook_feature")


def test_utf16_undeclared_worksheet_formula_is_rejected(tmp_path: Path) -> None:
    path = build_fixture(
        "relocated_formula_without_content_type", tmp_path / "utf16-formula.xlsx"
    )
    _rewrite_zip_member(
        path,
        "xl/fixture/worksheet.dat",
        lambda data: _move_formula_to_responsibility_field(data, "utf-16"),
    )

    _assert_rejected(path, "unsupported_workbook_feature")


def test_defined_name_formula_is_rejected(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "defined-name-formula.xlsx",
        [DEFAULT_HEADERS, ["192.0.2.1", 443, 443, "否", None]],
    )
    _rewrite_zip_member(
        path,
        "xl/workbook.xml",
        lambda data: data.replace(
            b"</workbook>",
            (
                b'<definedNames><definedName name="FixtureFormula">'
                b"OFFSET(Sheet!$A$1,0,0)</definedName></definedNames></workbook>"
            ),
        ),
    )

    _assert_rejected(path, "unsupported_workbook_feature")


def test_declared_dimensions_do_not_control_parser_iteration(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = _save_workbook(
        tmp_path / "inflated-dimensions.xlsx",
        [DEFAULT_HEADERS, ["192.0.2.1", 443, 443, "否", None]],
    )
    _rewrite_zip_member(
        path,
        "xl/worksheets/sheet1.xml",
        lambda data: re.sub(
            rb'<dimension ref="[^"]*"',
            b'<dimension ref="A1:XFD2000"',
            data,
            count=1,
        ),
    )

    original_iter_rows = ReadOnlyWorksheet.iter_rows
    observed_bounds: list[tuple[int | None, int | None]] = []

    def iter_rows_with_bounds(
        self: ReadOnlyWorksheet, *args: Any, **kwargs: Any
    ) -> Any:
        observed_bounds.append((kwargs.get("max_row"), kwargs.get("max_col")))
        return original_iter_rows(self, *args, **kwargs)

    monkeypatch.setattr(ReadOnlyWorksheet, "iter_rows", iter_rows_with_bounds)

    result = validate_customer_upload_workbook(path)

    assert result.record_count == 1
    assert observed_bounds == [(2, len(DEFAULT_HEADERS))]


def test_sparse_far_cell_is_rejected_before_parser_expansion(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "sparse-cell.xlsx",
        [DEFAULT_HEADERS, ["192.0.2.1", 443, 443, "否", None]],
    )
    _rewrite_zip_member(
        path,
        "xl/worksheets/sheet1.xml",
        lambda data: data.replace(
            b"</sheetData>",
            b'<row r="2000"><c r="XFD2000" t="n"><v>1</v></c></row></sheetData>',
        ),
    )

    _assert_rejected(path, "workbook_resource_limit")


def test_actual_file_size_over_twenty_mib_is_rejected_first(tmp_path: Path) -> None:
    path = tmp_path / "oversized.xlsx"
    with path.open("wb") as oversized:
        oversized.seek(MAX_WORKBOOK_BYTES)
        oversized.write(b"x")

    _assert_rejected(path, "upload_too_large")


def test_malformed_rejection_does_not_expose_path_or_parser_details(
    tmp_path: Path,
) -> None:
    path = tmp_path / "customer-secret-name.xlsx"
    path.write_bytes(b"not an XLSX: 192.0.2.123")

    with pytest.raises(CustomerUploadValidationError) as caught:
        validate_customer_upload_workbook(path)

    assert caught.value.__dict__ == {
        "code": "malformed_workbook",
        "field": None,
        "row": None,
    }
    assert caught.value.__cause__ is None
    assert caught.value.__context__ is None
    public_text = str(caught.value)
    assert "customer-secret-name" not in public_text
    assert "192.0.2.123" not in public_text
    assert "XLSX" not in public_text


def test_invalid_value_rejection_has_no_raw_exception_chain(tmp_path: Path) -> None:
    path = _save_workbook(
        tmp_path / "value.xlsx",
        [DEFAULT_HEADERS, ["customer-secret.example", 443, 443, "否", None]],
    )

    with pytest.raises(CustomerUploadValidationError) as caught:
        validate_customer_upload_workbook(path)

    assert caught.value.code == "invalid_required_value"
    assert caught.value.__cause__ is None
    assert caught.value.__context__ is None
