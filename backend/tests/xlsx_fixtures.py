"""Deterministic sanitized XLSX fixtures for validator tests."""

from __future__ import annotations

import re
import struct
import zlib
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Final
from zipfile import ZIP_DEFLATED, ZIP_STORED, ZipFile, ZipInfo

import openpyxl  # type: ignore[import-untyped]
from openpyxl import Workbook, load_workbook

from app.domain._xlsx_preflight import (
    MAX_ENTRY_UNCOMPRESSED_BYTES,
    MAX_TOTAL_UNCOMPRESSED_BYTES,
    MAX_ZIP_ENTRIES,
)

FIXED_TIME: Final = datetime(2026, 1, 1, tzinfo=UTC)
REL_NS: Final = "http://schemas.openxmlformats.org/package/2006/relationships"
MAIN_NS: Final = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS: Final = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)


def _set_reproducible_properties(workbook: Workbook) -> None:
    workbook.properties.creator = "Exposure-Agent fixture"
    workbook.properties.lastModifiedBy = "Exposure-Agent fixture"
    workbook.properties.created = FIXED_TIME.replace(tzinfo=None)
    workbook.properties.modified = FIXED_TIME.replace(tzinfo=None)


def _save_default_workbook(path: Path, *, formula: bool = False) -> None:
    workbook = Workbook()
    _set_reproducible_properties(workbook)
    worksheet = workbook.active
    worksheet.title = "资产清单"
    worksheet.append(
        [
            "资产IP",
            "起始端口",
            "结束端口",
            "是否web界面",
            "web界面url",
            "服务类型",
            "资产负责人",
            "资产所属部门",
            "端口负责人",
            "部门",
        ]
    )
    worksheet.append(
        [
            "192.0.2.10",
            443,
            443,
            "是",
            "https://fixture.example.invalid",
            "Fixture Service",
            "Example Owner",
            "Example Department",
            "Example Port Owner",
            "Example Operations",
        ]
    )
    worksheet.append(
        [
            "198.51.100.20",
            22,
            22,
            "否",
            None,
            "Fixture SSH",
            "Example Owner",
            "Example Department",
            "Example Port Owner",
            "Example Operations",
        ]
    )
    worksheet.append(
        [
            "203.0.113.30",
            8080,
            8080,
            "否",
            None,
            "Fixture API",
            "Example Owner",
            "Example Department",
            "Example Port Owner",
            "Example Operations",
        ]
    )
    if formula:
        worksheet["K2"] = "=1+1"
    workbook.save(path)
    workbook.close()
    _normalize_zip(path)


def _fixed_zip_info(name: str) -> ZipInfo:
    info = ZipInfo(name, date_time=(2026, 1, 1, 0, 0, 0))
    info.compress_type = ZIP_DEFLATED
    info.external_attr = 0o600 << 16
    return info


def _normalize_zip(path: Path) -> None:
    replacement = path.with_suffix(".normalized.xlsx")
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for source_info in source.infolist():
            source_data = source.read(source_info)
            if source_info.filename == "docProps/core.xml":
                source_data = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2026-01-01T00:00:00Z\g<2>",
                    source_data,
                )
            target.writestr(_fixed_zip_info(source_info.filename), source_data)
    replacement.replace(path)


def _append_declared_part(
    path: Path,
    *,
    name: str,
    data: bytes,
    content_type: str,
    relationships_name: str,
    relationship_type: str,
    relationship_target: str,
) -> None:
    replacement = path.with_suffix(".rewritten.xlsx")
    relationship_written = False
    declaration = (
        f'<Override PartName="/{name}" ContentType="{content_type}"/>'
    ).encode()
    relationship = (
        f'<Relationship Id="rIdFixture" Type="{relationship_type}" '
        f'Target="{relationship_target}"/>'
    ).encode()
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for source_info in source.infolist():
            source_data = source.read(source_info)
            if source_info.filename == "[Content_Types].xml":
                source_data = source_data.replace(
                    b"</Types>", declaration + b"</Types>"
                )
            elif source_info.filename == relationships_name:
                source_data = source_data.replace(
                    b"</Relationships>", relationship + b"</Relationships>"
                )
                relationship_written = True
            target.writestr(_fixed_zip_info(source_info.filename), source_data)
        if not relationship_written:
            target.writestr(
                _fixed_zip_info(relationships_name),
                (
                    f'<Relationships xmlns="{REL_NS}">'.encode()
                    + relationship
                    + b"</Relationships>"
                ),
            )
        target.writestr(_fixed_zip_info(name), data)
    replacement.replace(path)


def _add_vml_button(path: Path) -> None:
    _append_declared_part(
        path,
        name="xl/drawings/vmlDrawing1.vml",
        data=(
            b'<xml xmlns:v="urn:schemas-microsoft-com:vml" '
            b'xmlns:x="urn:schemas-microsoft-com:office:excel">'
            b'<v:shape id="_x0000_s1025"><x:ClientData ObjectType="Button">'
            b"<x:PrintObject>False</x:PrintObject>"
            b"</x:ClientData></v:shape></xml>"
        ),
        content_type="application/vnd.openxmlformats-officedocument.vmlDrawing",
        relationships_name="xl/worksheets/_rels/sheet1.xml.rels",
        relationship_type=f"{OFFICE_REL_NS}/vmlDrawing",
        relationship_target="../drawings/vmlDrawing1.vml",
    )
    replacement = path.with_suffix(".vml-button.xlsx")
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for source_info in source.infolist():
            source_data = source.read(source_info)
            if source_info.filename == "xl/worksheets/sheet1.xml":
                source_data = source_data.replace(
                    b"</worksheet>",
                    (
                        f'<legacyDrawing xmlns:r="{OFFICE_REL_NS}" '
                        'r:id="rIdFixture"/></worksheet>'
                    ).encode(),
                )
            target.writestr(_fixed_zip_info(source_info.filename), source_data)
    replacement.replace(path)


def _relocate_first_worksheet(path: Path) -> None:
    original_name = "xl/worksheets/sheet1.xml"
    relocated_name = "xl/fixture/worksheet.dat"
    replacement = path.with_suffix(".relocated.xlsx")
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        worksheet_data = source.read(original_name)
        for source_info in source.infolist():
            if source_info.filename == original_name:
                continue
            source_data = source.read(source_info)
            if source_info.filename == "[Content_Types].xml":
                source_data = source_data.replace(
                    b"/xl/worksheets/sheet1.xml", b"/xl/fixture/worksheet.dat"
                )
            elif source_info.filename == "xl/_rels/workbook.xml.rels":
                source_data = source_data.replace(
                    b'Target="/xl/worksheets/sheet1.xml"',
                    b'Target="/xl/fixture/worksheet.dat"',
                ).replace(
                    b'Target="worksheets/sheet1.xml"',
                    b'Target="fixture/worksheet.dat"',
                )
            target.writestr(_fixed_zip_info(source_info.filename), source_data)
        target.writestr(_fixed_zip_info(relocated_name), worksheet_data)
    replacement.replace(path)


def _add_chartsheet(path: Path, *, hidden: bool) -> None:
    _append_declared_part(
        path,
        name="xl/chartsheets/sheet1.xml",
        data=f'<chartsheet xmlns="{MAIN_NS}"/>'.encode(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.chartsheet+xml",
        relationships_name="xl/_rels/workbook.xml.rels",
        relationship_type=f"{OFFICE_REL_NS}/chartsheet",
        relationship_target="chartsheets/sheet1.xml",
    )
    replacement = path.with_suffix(".chartsheet.xlsx")
    state = b' state="hidden"' if hidden else b""
    chart_sheet = b'<sheet name="Chart" sheetId="2"' + state + b' r:id="rIdFixture"/>'
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for source_info in source.infolist():
            source_data = source.read(source_info)
            if source_info.filename == "xl/workbook.xml":
                source_data = source_data.replace(
                    b"</sheets>", chart_sheet + b"</sheets>"
                )
            target.writestr(_fixed_zip_info(source_info.filename), source_data)
    replacement.replace(path)


def _remove_content_type_override(path: Path, part_name: str) -> None:
    replacement = path.with_suffix(".without-content-type.xlsx")
    part_pattern = re.compile(
        rb'<Override PartName="/' + re.escape(part_name.encode()) + rb'"[^>]*/>'
    )
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for source_info in source.infolist():
            source_data = source.read(source_info)
            if source_info.filename == "[Content_Types].xml":
                source_data, replacements = part_pattern.subn(b"", source_data)
                if replacements != 1:
                    raise ValueError(
                        f"expected one Content Type override for {part_name}"
                    )
            target.writestr(_fixed_zip_info(source_info.filename), source_data)
    replacement.replace(path)


def _prefix_part(path: Path, name: str, prefix: bytes) -> None:
    replacement = path.with_suffix(".prefixed.xlsx")
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for source_info in source.infolist():
            source_data = source.read(source_info)
            if source_info.filename == name:
                source_data = prefix + source_data
            target.writestr(_fixed_zip_info(source_info.filename), source_data)
    replacement.replace(path)


def _create_png(width: int = 2_048, height: int = 3_300) -> bytes:
    signature = b"\x89PNG\r\n\x1a\n"

    def chunk(kind: bytes, data: bytes) -> bytes:
        return (
            struct.pack(">L", len(data))
            + kind
            + data
            + struct.pack(">L", zlib.crc32(kind + data) & 0xFFFFFFFF)
        )

    header = struct.pack(">2L5B", width, height, 8, 2, 0, 0, 0)
    pixel = bytes(range(256)) * ((width * 3 + 255) // 256)
    scanline = b"\x00" + pixel[: width * 3]
    image_data = zlib.compress(scanline * height, level=0)
    return (
        signature
        + chunk(b"IHDR", header)
        + chunk(b"IDAT", image_data)
        + chunk(b"IEND", b"")
    )


def _add_near_limit_static_image(path: Path) -> None:
    image = _create_png()
    replacement = path.with_suffix(".image.xlsx")
    with (
        ZipFile(path) as source,
        ZipFile(replacement, "w", compression=ZIP_DEFLATED, compresslevel=9) as target,
    ):
        for source_info in source.infolist():
            data = source.read(source_info)
            if source_info.filename == "[Content_Types].xml":
                data = data.replace(
                    b"</Types>",
                    b'<Default Extension="png" ContentType="image/png"/></Types>',
                )
            elif source_info.filename == "xl/worksheets/sheet1.xml":
                data = data.replace(
                    b"</worksheet>",
                    (
                        b'<drawing xmlns:r="http://schemas.openxmlformats.org/'
                        b'officeDocument/2006/relationships" r:id="rId1"/>'
                        b"</worksheet>"
                    ),
                )
            target.writestr(_fixed_zip_info(source_info.filename), data)
        target.writestr(
            _fixed_zip_info("xl/worksheets/_rels/sheet1.xml.rels"),
            (
                f'<Relationships xmlns="{REL_NS}"><Relationship Id="rId1" '
                f'Type="{OFFICE_REL_NS}/drawing" Target="../drawings/drawing1.xml"/>'
                "</Relationships>"
            ).encode(),
        )
        target.writestr(
            _fixed_zip_info("xl/drawings/drawing1.xml"),
            (
                '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
                'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"><xdr:oneCellAnchor>'
                "<xdr:from><xdr:col>11</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>1</xdr:row>"
                '<xdr:rowOff>0</xdr:rowOff></xdr:from><xdr:ext cx="9525" cy="9525"/>'
                '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="1" name="Boundary fixture"/>'
                "<xdr:cNvPicPr/></xdr:nvPicPr><xdr:blipFill><a:blip "
                f'xmlns:r="{OFFICE_REL_NS}" r:embed="rId1"/><a:stretch><a:fillRect/>'
                '</a:stretch></xdr:blipFill><xdr:spPr><a:prstGeom prst="rect"><a:avLst/>'
                "</a:prstGeom></xdr:spPr></xdr:pic><xdr:clientData/></xdr:oneCellAnchor></xdr:wsDr>"
            ).encode(),
        )
        target.writestr(
            _fixed_zip_info("xl/drawings/_rels/drawing1.xml.rels"),
            (
                f'<Relationships xmlns="{REL_NS}"><Relationship Id="rId1" '
                f'Type="{OFFICE_REL_NS}/image" Target="../media/boundary.png"/>'
                "</Relationships>"
            ).encode(),
        )
        image_info = _fixed_zip_info("xl/media/boundary.png")
        image_info.compress_type = ZIP_STORED
        target.writestr(image_info, image)
    replacement.replace(path)


def _stress_workbook_parts(
    row_count: int = 50_000, style_count: int = 64
) -> dict[str, bytes]:
    headers = [
        "资产IP",
        "起始端口",
        "结束端口",
        "是否web界面",
        "web界面url",
        "服务类型",
        "资产负责人",
        "资产所属部门",
        "端口负责人",
        "部门",
    ]
    addresses = (
        [f"192.0.2.{number}" for number in range(1, 255)]
        + [f"198.51.100.{number}" for number in range(1, 255)]
        + [f"203.0.113.{number}" for number in range(1, 255)]
    )
    ownership = [f"Example Responsibility {number:04d}" for number in range(1_024)]
    shared_strings = headers + addresses + ["否", "Fixture Service"] + ownership
    string_index = {value: index for index, value in enumerate(shared_strings)}
    shared_xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><sst xmlns="{MAIN_NS}" '
        f'count="{row_count * 8 + len(headers)}" uniqueCount="{len(shared_strings)}">'
        + "".join(f"<si><t>{value}</t></si>" for value in shared_strings)
        + "</sst>"
    )
    rows = [
        '<row r="1">'
        + "".join(
            f'<c r="{chr(65 + column)}1" t="s"><v>{string_index[header]}</v></c>'
            for column, header in enumerate(headers)
        )
        + "</row>"
    ]
    no_index = string_index["否"]
    service_index = string_index["Fixture Service"]
    for offset in range(row_count):
        row_number = offset + 2
        style = offset % style_count
        address_index = string_index[addresses[offset % len(addresses)]]
        owner_index = string_index[ownership[offset % len(ownership)]]
        rows.append(
            f'<row r="{row_number}">'
            f'<c r="A{row_number}" s="{style}" t="s"><v>{address_index}</v></c>'
            f'<c r="B{row_number}" s="{style}" t="n"><v>443</v></c>'
            f'<c r="C{row_number}" s="{style}" t="n"><v>443</v></c>'
            f'<c r="D{row_number}" s="{style}" t="s"><v>{no_index}</v></c>'
            f'<c r="F{row_number}" s="{style}" t="s"><v>{service_index}</v></c>'
            f'<c r="G{row_number}" s="{style}" t="s"><v>{owner_index}</v></c>'
            f'<c r="H{row_number}" s="{style}" t="s"><v>{owner_index}</v></c>'
            f'<c r="I{row_number}" s="{style}" t="s"><v>{owner_index}</v></c>'
            f'<c r="J{row_number}" s="{style}" t="s"><v>{owner_index}</v></c>'
            "</row>"
        )
    sheet_xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><worksheet xmlns="{MAIN_NS}">'
        f'<dimension ref="A1:J{row_count + 1}"/><sheetData>{"".join(rows)}</sheetData></worksheet>'
    )
    xfs = "".join(
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
        for _ in range(style_count)
    )
    styles_xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><styleSheet xmlns="{MAIN_NS}">'
        '<fonts count="1"><font><name val="Arial"/><sz val="10"/></font></fonts>'
        '<fills count="2"><fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="gray125"/></fill></fills>'
        '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        f'<cellXfs count="{style_count}">{xfs}</cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        "</styleSheet>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        "</Types>"
    )
    root_rels = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="{REL_NS}">'
        f'<Relationship Id="rId1" Type="{OFFICE_REL_NS}/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    workbook_xml = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="{MAIN_NS}" '
        f'xmlns:r="{OFFICE_REL_NS}"><sheets><sheet name="资产清单" sheetId="1" r:id="rId1"/>'
        "</sheets></workbook>"
    )
    workbook_rels = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="{REL_NS}">'
        f'<Relationship Id="rId1" Type="{OFFICE_REL_NS}/worksheet" Target="worksheets/sheet1.xml"/>'
        f'<Relationship Id="rId2" Type="{OFFICE_REL_NS}/styles" Target="styles.xml"/>'
        f'<Relationship Id="rId3" Type="{OFFICE_REL_NS}/sharedStrings" Target="sharedStrings.xml"/>'
        "</Relationships>"
    )
    return {
        "[Content_Types].xml": content_types.encode(),
        "_rels/.rels": root_rels.encode(),
        "xl/workbook.xml": workbook_xml.encode(),
        "xl/_rels/workbook.xml.rels": workbook_rels.encode(),
        "xl/styles.xml": styles_xml.encode(),
        "xl/sharedStrings.xml": shared_xml.encode(),
        "xl/worksheets/sheet1.xml": sheet_xml.encode(),
    }


def _write_parts(path: Path, parts: dict[str, bytes]) -> None:
    with ZipFile(path, "w", compression=ZIP_DEFLATED, compresslevel=9) as archive:
        for name, data in parts.items():
            archive.writestr(_fixed_zip_info(name), data)


def _write_repeated_member(member: Any, size: int, byte: bytes) -> None:
    block = byte * (1024 * 1024)
    remaining = size
    while remaining:
        chunk_size = min(remaining, len(block))
        member.write(block[:chunk_size])
        remaining -= chunk_size


def build_xlsx_fixture(name: str, path: Path) -> Path:
    """Build one deterministic, sanitized fixture at a caller-owned path."""
    path.parent.mkdir(parents=True, exist_ok=True)
    if name == "row_shared_style_boundary":
        _write_parts(path, _stress_workbook_parts())
        return path
    if name == "entry_count_bomb":
        with ZipFile(path, "w", compression=ZIP_DEFLATED, compresslevel=9) as archive:
            for index in range(MAX_ZIP_ENTRIES + 1):
                archive.writestr(_fixed_zip_info(f"fixture/{index:04d}.bin"), b"")
        return path

    _save_default_workbook(
        path,
        formula=name
        in {
            "formula",
            "relocated_formula",
            "relocated_formula_without_content_type",
            "padded_relocated_formula",
        },
    )
    if name in {"default_v1", "formula"}:
        return path
    if name == "relocated_formula":
        _relocate_first_worksheet(path)
        return path
    if name == "relocated_formula_without_content_type":
        _relocate_first_worksheet(path)
        _remove_content_type_override(path, "xl/fixture/worksheet.dat")
        return path
    if name == "padded_relocated_formula":
        _relocate_first_worksheet(path)
        _remove_content_type_override(path, "xl/fixture/worksheet.dat")
        _prefix_part(path, "xl/fixture/worksheet.dat", b" " * 5_000)
        return path
    if name == "near_request_limit":
        _add_near_limit_static_image(path)
    elif name == "conditional_format_formula":
        workbook = load_workbook(path)
        worksheet = workbook.active
        worksheet.conditional_formatting.add(
            "A2:A4",
            openpyxl.formatting.rule.FormulaRule(formula=["A2=1"]),
        )
        _set_reproducible_properties(workbook)
        workbook.save(path)
        workbook.close()
        _normalize_zip(path)
    elif name == "data_validation_formula":
        workbook = load_workbook(path)
        worksheet = workbook.active
        validation = openpyxl.worksheet.datavalidation.DataValidation(
            type="whole", operator="between", formula1="1", formula2="9"
        )
        worksheet.add_data_validation(validation)
        validation.add("A2:A4")
        _set_reproducible_properties(workbook)
        workbook.save(path)
        workbook.close()
        _normalize_zip(path)
    elif name == "table_formula":
        _append_declared_part(
            path,
            name="xl/tables/table1.xml",
            data=(
                b"<table><tableColumns><tableColumn>"
                b"<calculatedColumnFormula>A1+1</calculatedColumnFormula>"
                b"<totalsRowFormula>SUM(A:A)</totalsRowFormula>"
                b"</tableColumn></tableColumns></table>"
            ),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml"
            ),
            relationships_name="xl/worksheets/_rels/sheet1.xml.rels",
            relationship_type=f"{OFFICE_REL_NS}/table",
            relationship_target="../tables/table1.xml",
        )
    elif name == "hidden_sheet":
        workbook = load_workbook(path)
        hidden = workbook.create_sheet("Hidden fixture")
        hidden.sheet_state = "hidden"
        _set_reproducible_properties(workbook)
        workbook.save(path)
        workbook.close()
        _normalize_zip(path)
    elif name == "hidden_chartsheet":
        _add_chartsheet(path, hidden=True)
    elif name == "visible_chartsheet":
        _add_chartsheet(path, hidden=False)
    elif name == "external_link":
        _append_declared_part(
            path,
            name="xl/fixture/external-link.xml",
            data=f'<externalLink xmlns="{MAIN_NS}"/>'.encode(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml",
            relationships_name="xl/_rels/workbook.xml.rels",
            relationship_type=f"{OFFICE_REL_NS}/externalLink",
            relationship_target="fixture/external-link.xml",
        )
    elif name == "data_connection":
        _append_declared_part(
            path,
            name="xl/fixture/connections.xml",
            data=f'<connections xmlns="{MAIN_NS}" count="0"/>'.encode(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml",
            relationships_name="xl/_rels/workbook.xml.rels",
            relationship_type=f"{OFFICE_REL_NS}/connections",
            relationship_target="fixture/connections.xml",
        )
    elif name == "embedded_object":
        _append_declared_part(
            path,
            name="xl/fixture/object.bin",
            data=b"fixture-active-object",
            content_type="application/vnd.openxmlformats-officedocument.oleObject",
            relationships_name="xl/worksheets/_rels/sheet1.xml.rels",
            relationship_type=f"{OFFICE_REL_NS}/oleObject",
            relationship_target="../fixture/object.bin",
        )
    elif name == "vml_button":
        _add_vml_button(path)
    elif name == "compression_bomb":
        with ZipFile(path, "a", compression=ZIP_DEFLATED, compresslevel=9) as archive:
            info = _fixed_zip_info("xl/sharedStrings.xml")
            with archive.open(info, "w", force_zip64=False) as member:
                _write_repeated_member(member, MAX_ENTRY_UNCOMPRESSED_BYTES + 1, b"A")
    elif name == "total_size_bomb":
        member_size = MAX_TOTAL_UNCOMPRESSED_BYTES // 5 + 1
        with ZipFile(path, "a", compression=ZIP_DEFLATED, compresslevel=9) as archive:
            for index in range(5):
                info = _fixed_zip_info(f"fixture/total-{index}.bin")
                with archive.open(info, "w", force_zip64=False) as member:
                    _write_repeated_member(member, member_size, b"B")
    else:
        raise ValueError(f"unknown fixture kind: {name}")
    return path
